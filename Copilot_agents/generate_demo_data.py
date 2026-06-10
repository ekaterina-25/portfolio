"""Generates synthetic demo Excel files for both Copilot agent portfolio demos."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE = Path(__file__).parent

# ── item_data_normalizer: demo_input.xlsx ─────────────────────────────────

input_headers = [
    "Item Code", "Name", "Description (English)", "Description (Finnish)",
    "Specification", "Standard",
]

# INT-2019 replaces the real company-internal standard.
# Rows intentionally contain typical data quality issues:
# all-caps names, mixed DIN+ISO standards, comma decimals,
# double spaces, missing standards, and wrong case in Finnish.
input_rows = [
    ["SYN-00001", "ADJUSTING SCREW M16 x350",        "Adjusting screw",           "Säätöruuvi",             "M16x350",               "INT-2019"            ],
    ["SYN-00002", "ADJUSTING SCREW M10 x 220",       "Adjusting screw",           "Säätöruuvi",             "M10 x220",              "INT-2019"            ],
    ["SYN-00003", "Anchor bolt S-KA 12/80 L178",     "Anchor bolt",               "Kiila-ankkuri",          "S-KA 12/80 L178",       None                  ],
    ["SYN-00004", "CARRIAGE BOLT",                   "Carriage bolt",             "Lukkoruuvi",             "M10 x100  8.8 Zn",      "DIN 603, EN ISO 8678"],
    ["SYN-00005", "CUP SQUARE BOLT M12 x 50 4.6",   "Cup square bolt",           "Lukkoruuvi",             "M12x50 g4.6",           "DIN 603, EN ISO 8677"],
    ["SYN-00006", "SOCKET HEAD SCREW M8x80 DIN912",  "Hex socket screw",          "Kuusiokoloruuvi",        "M8x80 - 8.8",           "EN ISO 4762, DIN 912"],
    ["SYN-00007", "HEX. SOCKET HEAD SCREW",          "Hex socket screw",          "KUUSIOKOLORUUVI",        "M6 x 1,0 x 25 8.8 ZN", "ISO 4762"            ],
    ["SYN-00008", "HEXAGON SCREW",                   "Hexagon screw",             "Kuusioruuvi",            "M24x120 8.8 Zn",        "DIN 931, ISO 4014"   ],
    ["SYN-00009", "Countersunk socket screw M5x20",  "Countersunk socket screw",  "Uppokantaruuvi",         "M5x20 10.9",            "ISO 10642"           ],
    ["SYN-00010", "SOCKET HEAD SCREW A4-70",         "Socket head cap screw",     "Kuusiokoloruuvi",        "M3x12 A4-70",           "ISO 4762"            ],
    ["SYN-00011", "HEXAGON NUT",                     "Hex nut",                   "KUUSIOMUTTERI",          "M16 8",                 "DIN 934, ISO 4032"   ],
    ["SYN-00012", "Self-locking nut  M12",           "Self-locking nut",          "Jousimutteri",           "M12  8",                "DIN 985, ISO 7042"   ],
    ["SYN-00013", "PLAIN WASHER",                    "Plain washer",              "Aluslevy",               "10",                    "DIN 125, ISO 7089"   ],
    ["SYN-00014", "SIIPIRUUVI M6 x16",               "Wing screw",                "SIIPIRUUVI",             "M6x16",                 "DIN 316"             ],
    ["SYN-00015", "Hex socket set screw M6x8 10.9",  "Hex socket set screw",      "Kuusiokolopidätinruuvi", "M6x8 10.9",             "DIN 916, ISO 4029"   ],
]

wb_in = openpyxl.Workbook()
ws = wb_in.active
ws.title = "Data"
h_fill = PatternFill("solid", fgColor="4F81BD")
h_font = Font(bold=True, color="FFFFFF")
ws.append(input_headers)
for cell in ws[1]:
    cell.fill = h_fill
    cell.font = h_font
    cell.alignment = Alignment(horizontal="center")
for row in input_rows:
    ws.append(row)
col_widths = [14, 38, 28, 28, 26, 26]
for col_idx, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.freeze_panes = "B2"
out_path = BASE / "item_data_normalizer" / "demo_data" / "demo_input.xlsx"
wb_in.save(out_path)
print(f"Created {out_path}  ({len(input_rows)} rows)")


# ── item_data_normalizer: demo_output.xlsx ────────────────────────────────

output_headers = ["Item Code", "Name", "Description (English)", "Description (Finnish)", "Specification"]

output_rows = [
    ["SYN-00001", "Adjusting screw INT-2019 M16×350",                           "Adjusting screw",                       "Säätöruuvi",                "INT-2019 M16×350"      ],
    ["SYN-00002", "Adjusting screw INT-2019 M10×220",                           "Adjusting screw",                       "Säätöruuvi",                "INT-2019 M10×220"      ],
    ["SYN-00003", "Anchor bolt S-KA 12/80 L178",                                "Anchor bolt",                           "Kiila-ankkuri",             "S-KA 12/80 L178"       ],
    ["SYN-00004", "Cup square bolt ISO 8678 Zn 8.8 M10×100",                    "Cup square bolt",                       "Lukkoruuvi",                "ISO 8678 Zn 8.8 M10×100"],
    ["SYN-00005", "Cup square bolt ISO 8677 4.6 M12×50",                        "Cup square bolt",                       "Lukkoruuvi",                "ISO 8677 4.6 M12×50"   ],
    ["SYN-00006", "Hexagon socket head cap screw ISO 4762 8.8 M8×80",           "Hexagon socket head cap screw",         "Kuusiokoloruuvi",           "ISO 4762 8.8 M8×80"    ],
    ["SYN-00007", "Hexagon socket head cap screw ISO 4762 Zn 8.8 M6×25",        "Hexagon socket head cap screw",         "Kuusiokoloruuvi",           "ISO 4762 Zn 8.8 M6×25" ],
    ["SYN-00008", "Hexagon head screw ISO 4014 Zn 8.8 M24×120",                 "Hexagon head screw",                    "Kuusioruuvi",               "ISO 4014 Zn 8.8 M24×120"],
    ["SYN-00009", "Hexagon socket countersunk head screw ISO 10642 10.9 M5×20", "Hexagon socket countersunk head screw", "Kuusiokoloruuvi uppokanta", "ISO 10642 10.9 M5×20"  ],
    ["SYN-00010", "Hexagon socket head cap screw ISO 4762 A4-70 M3×12",         "Hexagon socket head cap screw",         "Kuusiokoloruuvi",           "ISO 4762 A4-70 M3×12"  ],
    ["SYN-00011", "Hexagon nut ISO 4032 8 M16",                                  "Hexagon nut",                           "Kuusiomutteri",             "ISO 4032 8 M16"        ],
    ["SYN-00012", "Self-locking nut ISO 7042 8 M12",                             "Self-locking nut",                      "Jousimutteri",              "ISO 7042 8 M12"        ],
    ["SYN-00013", "Plain washer ISO 7089 M10",                                   "Plain washer",                          "Tasainen aluslevy",         "ISO 7089 M10"          ],
    ["SYN-00014", "Wing screw DIN 316 M6×16",                                    "Wing screw",                            "Siipiruuvi",                "DIN 316 M6×16"         ],
    ["SYN-00015", "Hexagon socket set screw ISO 4029 10.9 M6×8",                "Hexagon socket set screw",              "Kuusiokolopidätinruuvi",    "ISO 4029 10.9 M6×8"    ],
]

wb_out = openpyxl.Workbook()
ws2 = wb_out.active
ws2.title = "Result"
r_fill = PatternFill("solid", fgColor="375623")
r_font = Font(bold=True, color="FFFFFF")
ws2.append(output_headers)
for cell in ws2[1]:
    cell.fill = r_fill
    cell.font = r_font
    cell.alignment = Alignment(horizontal="center")
for row in output_rows:
    ws2.append(row)
for i, w in enumerate([14, 58, 44, 32, 42], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "B2"
out_path2 = BASE / "item_data_normalizer" / "demo_data" / "demo_output.xlsx"
wb_out.save(out_path2)
print(f"Created {out_path2}  ({len(output_rows)} rows)")


# ── product_classifier: material_groups_demo.xlsx ─────────────────────────

groups = [
    ("31100", "Screws"),
    ("31200", "Nuts"),
    ("31300", "Washers"),
    ("31400", "Pins, Retaining Rings, Circlips"),
    ("31800", "Springs, Dampers, Bumpers"),
    ("32200", "Roller Bearings"),
    ("32900", "Linear Elements"),
    ("33200", "Couplings, Clutches"),
    ("35100", "Gear Motors"),
    ("35200", "Motors"),
    ("35400", "HVAC and Plumbing, Pumps"),
    ("35600", "HVAC and Plumbing, Valves, Strainers"),
    ("35650", "HVAC and Plumbing, Filters"),
    ("36000", "Hydraulic Cylinders"),
    ("36100", "Hydraulic Valves"),
    ("36500", "Hydraulic Motors, Pumps"),
    ("36800", "Hydraulic Filters"),
    ("37000", "Pneumatic Cylinders"),
    ("37100", "Pneumatic Valves"),
    ("37600", "Pneumatic Filters"),
    ("37900", "Gauges, Sensors"),
    ("38500", "Sensors, Encoders"),
    ("39200", "Sealings"),
    ("91000", "Installation Accessory"),
    ("96800", "Inverters, Frequency Converters"),
]

wb_mg = openpyxl.Workbook()
ws3 = wb_mg.active
ws3.title = "Material Groups"
mg_fill = PatternFill("solid", fgColor="4F81BD")
mg_font = Font(bold=True, color="FFFFFF")
ws3.append(["Group Code", "Group Name"])
for cell in ws3[1]:
    cell.fill = mg_fill
    cell.font = mg_font
    cell.alignment = Alignment(horizontal="center")
for code, name in groups:
    ws3.append([code, name])
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 42
ws3.freeze_panes = "A2"
out_path3 = BASE / "product_classifier" / "demo_data" / "material_groups_demo.xlsx"
wb_mg.save(out_path3)
print(f"Created {out_path3}  ({len(groups)} groups)")
