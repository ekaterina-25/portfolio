import io
import os
import re
import sys

import pandas as pd
import streamlit as st

# Ensure validate_items.py is always loaded from this directory,
# not from any other location on the Python path.
sys.path.insert(0, os.path.dirname(__file__))

from validate_items import (
    CHECK_COLS, CHECK_LABELS, CODE_LIMIT, DESC_LIMIT, FORBIDDEN_SYMBOLS,
    UPPERCASE_COLS, auto_fix, bad_data_cells, is_pass,
    load_reference, manual_fix_rows, run_checks, suggest_fixes, summary,
)

DATA_DIR   = os.path.join(os.path.dirname(__file__), "data")
DEMO_ITEMS = os.path.join(DATA_DIR, "validation_data.xlsx")
DEMO_REF   = os.path.join(DATA_DIR, "reference_data.xlsx")

# Bootstrap alert palette
PASS_COLOR  = "#d4edda"
FAIL_COLOR  = "#f8d7da"
FIXED_COLOR = "#c3e6cb"   # slightly deeper green — marks auto-fixed cells
WARN_COLOR  = "#fff3cd"   # yellow — reserved for suggestion cells

# Display labels with fix-type symbols — UI only, not written to Excel.
#   🔧 = auto-fixable now   💡 = suggestion   ✏️ = manual fix
DISPLAY_LABELS = {
    "check_product_group": "💡 Product group valid",
    "check_basic_name":    "💡 Basic name valid",
    "check_symbols":       "✏️ No forbidden symbols",
    "check_uppercase":     "🔧 Uppercase",
    "check_spaces":        "🔧 No extra spaces",
    "check_length":        "✏️ Field length",
    "check_name_in_desc":  "✏️ Basic name in description",
    "check_spec_in_desc":  "✏️ Specification in all descriptions",
}

# ---------------------------------------------------------------------------
# Helper — build the combined corrections table from current item data.
# Called on first run and again after each Apply so the table always
# reflects the current state of the data (resolved errors drop out,
# stale "before" values are updated to match the corrected data).
# ---------------------------------------------------------------------------
def _build_corrections(items: pd.DataFrame, results: pd.DataFrame, ref: pd.DataFrame):
    _, auto_changes = auto_fix(items)
    suggestions     = suggest_fixes(items, ref)
    manual_df       = manual_fix_rows(results)

    auto_rows   = auto_changes.assign(Accept=True).copy()
    sugg_rows   = suggestions.assign(Accept=False).copy()
    manual_rows = manual_df.assign(Accept=False).copy()

    combined = pd.concat([auto_rows, sugg_rows, manual_rows], ignore_index=True)
    combined["error_description"] = combined.get("error_description", pd.Series(dtype=str))
    combined["error_description"] = combined["error_description"].fillna("")

    combined["Fix type"] = combined["fix_type"].map({
        "uppercase":  "🔧 Auto-fix",
        "spaces":     "🔧 Auto-fix",
        "suggestion": "💡 Suggestion",
        "manual":     "✏️ Manual",
    })

    def _err_desc(row):
        if row["fix_type"] == "manual":
            return row["error_description"] or "Manual fix required"
        if row["fix_type"] == "uppercase":
            return "Contains lowercase letters"
        if row["fix_type"] == "spaces":
            return "Leading, trailing, or extra spaces"
        if row["column"] == "basic_name":
            return "Unknown basic name — not in reference"
        if row["column"] == "product_group":
            return "Unknown product group code"
        return "Invalid value"

    combined["Error description"] = combined.apply(_err_desc, axis=1)

    desc_map = items.set_index("item_id")["description_en"].to_dict() \
               if "item_id" in items.columns and "description_en" in items.columns else {}
    combined["desc_en"] = combined["item_id"].map(desc_map)

    corrections_table = combined[
        ["Accept", "Fix type", "item_id", "column",
         "Error description", "before", "after", "desc_en"]
    ].rename(columns={
        "item_id": "Item ID",
        "column":  "Column",
        "before":  "Wrong data",
        "after":   "Suggested fix",
        "desc_en": "Description EN (ref)",
    })

    return corrections_table, suggestions


# ---------------------------------------------------------------------------
# Validation helper — checks one accepted fix value against applicable rules.
# fix_type scopes which rules apply: auto-fix rows target exactly one
# problem (uppercase or spaces) and should not be blocked by pre-existing
# violations that have their own separate fix rows.
# ---------------------------------------------------------------------------
def _validate_fix(item_id, col, val, valid_names, valid_groups, fix_type="manual"):
    # Controlled-vocabulary checks — always apply regardless of fix type.
    if col == "basic_name" and val not in valid_names:
        return f"• Item **{item_id}**: '{val}' is not a valid basic name"
    if col == "product_group" and val not in valid_groups:
        return f"• Item **{item_id}**: '{val}' is not a valid product group code"

    # Uppercase required for description fields.
    # basic_name is already caught by the reference check above.
    if col in UPPERCASE_COLS and col != "basic_name" and any(c.islower() for c in val):
        return f"• Item **{item_id}** ({col}): value must be uppercase"

    # Length limits — skip for auto-fix rows.
    # Uppercase conversion never changes length; spaces normalisation never
    # increases it. A pre-existing length violation has its own ✏️ fix row.
    if fix_type != "auto":
        if col in ("description_en", "description_fi", "description_de") and len(val) > DESC_LIMIT:
            return f"• Item **{item_id}** ({col}): too long — {len(val)} chars (max {DESC_LIMIT})"
        if col == "product_code" and len(val) > CODE_LIMIT:
            return f"• Item **{item_id}** (product_code): too long — {len(val)} chars (max {CODE_LIMIT})"

    # Forbidden symbols
    bad_syms = [s for s in FORBIDDEN_SYMBOLS if s in val]
    if bad_syms:
        return (f"• Item **{item_id}** ({col}): contains forbidden symbol(s): "
                + ", ".join(repr(s) for s in bad_syms))

    # Extra whitespace
    if re.search(r"^\s+|\s+$|\s{2,}", val):
        return f"• Item **{item_id}** ({col}): leading, trailing, or extra spaces"

    return None


# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Item Data Validation", layout="wide")

# Make tab labels more legible — default Streamlit tab text is too small.
st.markdown("""
<style>
/* Tab button container */
button[data-baseweb="tab"] {
    font-size: 1.3rem !important;
    font-weight: 600 !important;
    padding: 0.5rem 1.4rem !important;
}
/* Streamlit wraps tab text in a <p> in newer versions */
button[data-baseweb="tab"] p,
button[data-baseweb="tab"] span {
    font-size: 1.3rem !important;
    font-weight: 600 !important;
}
/* Active tab: blue underline */
button[data-baseweb="tab"][aria-selected="true"] {
    color: #0d6efd !important;
    border-bottom: 3px solid #0d6efd !important;
}
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------
with st.sidebar:
    st.title("Data sources")

    data_source = st.radio(
        "Data source", ["Demo data", "Upload own data"],
        index=0, horizontal=True,
    )

    if data_source == "Demo data":
        items_file = DEMO_ITEMS
        ref_file   = DEMO_REF
    else:
        items_file = st.file_uploader("Item data (.xlsx)", type="xlsx")
        ref_file   = st.file_uploader("Reference data (.xlsx)", type="xlsx")

    st.divider()
    st.markdown(
        "**Reference data** must contain columns:\\* "
        "`product_group_id`, `description_en`, `basic_name`."
    )
    st.markdown(
        "**Item data** must contain columns:\\* "
        "`item_id`, `product_group`, `basic_name`, `specification`, "
        "`description_en`, `description_fi`, `description_de`, `product_code`, `status`."
    )
    st.caption(
        "\\* Column names are fixed in this demo version. "
        "A production implementation would support configurable column mapping."
    )
    st.divider()
    run_button = st.button("▶ Run validation", type="primary", use_container_width=True)

# ---------------------------------------------------------------------------
# Main — title and description always visible
# ---------------------------------------------------------------------------
st.title("Item Data Validation")
st.markdown(
    "Example of automated data validation against item master data standardisation rules. "
    "Each item is checked across four categories:"
)
st.markdown(
    "- **Reference catalogue** — product group codes and basic names must match approved values\n"
    "- **Field content** — forbidden symbols, uppercase convention, no extra spaces\n"
    "- **Field length** — ERP import limits (descriptions max 40 chars, part numbers max 30)\n"
    "- **Cross-field consistency** — specification and basic name must appear in all language descriptions"
)

with st.expander("How to use", expanded=True):
    st.markdown(
        "1. **Select a data source** in the sidebar — use demo data (recommended) or upload your own files.\n"
        "2. **Run validation** — click *Run validation* to check all items against the rules.\n"
        "3. **Review results** — the *📊 Results* tab shows which items and checks failed.\n"
        "4. **Fix errors** — go to the *🔧 Error fix* tab:\n"
        "   - 🔧 Auto-fix rows are pre-selected; accept, update or deselect if needed.\n"
        "   - 💡 Suggestion rows show the closest match from the reference — check before accepting.\n"
        "   - ✏️ Manual rows have no suggestion — you may type the corrected value yourself.\n"
        "   - Click **Apply selected fixes** to apply accepted corrections.\n"
        "5. **Download** — export results and corrected data from the *⬇ Download* tab."
    )

# Clear all state when data source changes so stale results don't persist
_sig = data_source + str(getattr(items_file, "name", items_file))
if st.session_state.get("val_sig") != _sig:
    for _k in ["val_results", "val_summary", "val_items", "val_ref",
               "val_fixed_applied", "val_fixed_items", "val_fixed_results",
               "val_fixed_summary", "val_accepted_cells", "val_rejected_cells",
               "editor_key_version"]:
        st.session_state.pop(_k, None)
    st.session_state["val_sig"] = _sig

if not run_button and "val_results" not in st.session_state:
    st.stop()

# ---------------------------------------------------------------------------
# Run validation
# ---------------------------------------------------------------------------
if run_button:
    if not items_file or not ref_file:
        st.warning("Select a data source in the sidebar before running.")
        st.stop()
    try:
        items = pd.read_excel(items_file)
        ref   = load_reference(ref_file)
    except Exception as e:
        st.error(f"Could not read files: {e}")
        st.stop()

    with st.spinner("Running validation checks…"):
        results = run_checks(items, ref)
        summ    = summary(results)

    corrections_table, suggestions = _build_corrections(items, results, ref)

    # Clear any previous fix state when re-running
    for _k in ["val_fixed_applied", "val_fixed_items", "val_fixed_results",
               "val_fixed_summary", "val_accepted_cells", "val_rejected_cells"]:
        st.session_state.pop(_k, None)

    st.session_state.update(
        val_results=results,
        val_summary=summ,
        val_items=items,
        val_ref=ref,
        val_corrections=corrections_table,
        val_suggestions=suggestions,
    )

if "val_results" not in st.session_state:
    st.stop()

# Load from session state — use fixed data in all views if corrections have been applied
results = st.session_state["val_results"]
summ    = st.session_state["val_summary"]
items   = st.session_state["val_items"]
ref     = st.session_state["val_ref"]

fixed_applied  = st.session_state.get("val_fixed_applied", False)
active_results = st.session_state.get("val_fixed_results", results) if fixed_applied else results
active_items   = st.session_state.get("val_fixed_items",   items)   if fixed_applied else items

# ---------------------------------------------------------------------------
# KPI row — reflects fixed data if corrections have been applied
# ---------------------------------------------------------------------------
total      = len(active_results)
any_fail   = active_results.apply(
    lambda row: any(not is_pass(c, row[c]) for c in CHECK_COLS), axis=1
)
items_fail = any_fail.sum()
pass_rate  = (total - items_fail) / total

c1, c2, c3 = st.columns(3)
c1.metric("Total items",       total)
c2.metric("Items with errors", items_fail)
c3.metric("Overall pass rate", f"{pass_rate:.0%}")
if fixed_applied:
    orig_fail = st.session_state["val_results"].apply(
        lambda row: any(not is_pass(c, row[c]) for c in CHECK_COLS), axis=1
    ).sum()
    st.caption(f"✅ Corrections applied — errors reduced from {orig_fail} to {items_fail}")

# ---------------------------------------------------------------------------
# Tabs
# ---------------------------------------------------------------------------
tab_results, tab_autofix, tab_download = st.tabs(
    ["📊 Results", "🔧 Error fix", "⬇ Download"]
)

# ── Shared styling helpers ────────────────────────────────────────────────
def colour_pass_rate(val):
    """Green = 100 %, yellow = 90–99 %, red < 90 %."""
    pct = int(val.strip("%"))
    if pct == 100:
        return f"background-color: {PASS_COLOR}"
    if pct >= 90:
        return f"background-color: {WARN_COLOR}"
    return f"background-color: {FAIL_COLOR}"

def colour_cell(val):
    """Colour check-result cells: bool or ok/error string."""
    if isinstance(val, bool):
        return f"background-color: {PASS_COLOR if val else FAIL_COLOR}"
    s = str(val).strip().lower()
    if s in ("ok", "true"):
        return f"background-color: {PASS_COLOR}"
    if s == "false":
        return f"background-color: {FAIL_COLOR}"
    if s == "specification empty":
        return f"background-color: {WARN_COLOR}"
    if s:
        return f"background-color: {FAIL_COLOR}"
    return ""

# ── Tab 1: Results ────────────────────────────────────────────────────────
with tab_results:

    # ── Check summary ─────────────────────────────────────────────────────
    active_summ  = st.session_state["val_fixed_summary"] if fixed_applied else summ
    summ_display = active_summ.copy()
    summ_display["Check"] = summ_display["Check"].map(
        {v: DISPLAY_LABELS[k] for k, v in CHECK_LABELS.items()}
    )
    if fixed_applied:
        summ_display["Errors fixed"] = (summ["Fail"].values - active_summ["Fail"].values)
    styled_summ = summ_display.style.map(colour_pass_rate, subset=["Pass rate"])
    st.dataframe(styled_summ, use_container_width=True, hide_index=True)
    st.caption("🔧 auto-fixable  •  💡 suggestion planned  •  ✏️ manual fix required")

    st.divider()

    # ── Item-level results ────────────────────────────────────────────────
    if fixed_applied:
        st.info("Showing data after corrections. Fixed cells are highlighted green.")

    filter_option = st.radio(
        "Show", ["All items", "Items with errors only"], horizontal=True,
    )
    selected_checks = st.multiselect(
        "Filter by failed check",
        options=list(DISPLAY_LABELS.values()),
        default=[], placeholder="All checks",
    )

    view = active_results.copy()

    if filter_option == "Items with errors only":
        view = view[view.apply(
            lambda row: any(not is_pass(c, row[c]) for c in CHECK_COLS), axis=1
        )]

    if selected_checks:
        col_map = {v: k for k, v in DISPLAY_LABELS.items()}
        selected_cols = [col_map[s] for s in selected_checks]
        view = view[view.apply(
            lambda row: any(not is_pass(c, row[c]) for c in selected_cols), axis=1
        )]

    st.caption(f"Showing {len(view)} of {total} items")

    display_cols = ["item_id", "product_group", "basic_name", "specification",
                    "description_en", "description_fi", "description_de",
                    "product_code", "status"]
    display_cols = [c for c in display_cols if c in view.columns]
    check_display = view[display_cols + CHECK_COLS].rename(columns=DISPLAY_LABELS)

    bad    = bad_data_cells(view)
    # fixed_cells  → green  (accepted corrections)
    # suggested_cells → yellow (suggestion available, not yet acted on)
    fixed_cells     = set()
    suggested_cells = set()
    rejected_cells  = set()
    if fixed_applied:
        fixed_cells    = st.session_state.get("val_accepted_cells", set())
        rejected_cells = st.session_state.get("val_rejected_cells", set())
    elif "val_suggestions" in st.session_state:
        # Before any Apply: show pending suggestion cells in yellow
        sdf = st.session_state["val_suggestions"]
        suggested_cells = set(zip(sdf["item_id"], sdf["column"]))

    def _colour_data_cells(df):
        result = pd.DataFrame("", index=df.index, columns=df.columns)
        for col in display_cols:
            if col not in bad.columns:
                continue
            for idx in df.index:
                item_id = view.at[idx, "item_id"] if "item_id" in view.columns else idx
                if (item_id, col) in fixed_cells:
                    result.at[idx, col] = f"background-color: {FIXED_COLOR}"
                elif (item_id, col) in rejected_cells or (item_id, col) in suggested_cells:
                    result.at[idx, col] = f"background-color: {WARN_COLOR}"
                elif bad.at[idx, col]:
                    result.at[idx, col] = f"background-color: {FAIL_COLOR}"
        return result

    check_label_cols = list(DISPLAY_LABELS.values())
    styled_detail = (
        check_display.style
        .apply(_colour_data_cells, axis=None)
        .map(colour_cell, subset=check_label_cols)
    )
    st.dataframe(styled_detail, use_container_width=True, hide_index=True)


# ── Tab 2: Error fix ──────────────────────────────────────────────────────
def render_fix_tab():
    st.markdown(
        "Review and edit corrections before applying. "
        "🔧 rows are pre-selected (deterministic fixes). "
        "💡 rows are suggestions — check them to accept, or edit the value first. "
        "✏️ rows require you to type a corrected value."
    )
    st.caption("🔧 Auto-fix = uppercase / spaces  •  💡 Suggestion = closest match from reference  •  ✏️ Manual fix required")

    if "val_corrections" not in st.session_state:
        st.info("Run validation first.")
        return

    # Compute key before buttons so we can read the current editor state in
    # Select All / Deselect All (to preserve any text the user has already typed).
    _editor_version = st.session_state.get("editor_key_version", 0)
    _editor_key     = f"corrections_editor_{_editor_version}"

    btn1, btn2, btn3, _ = st.columns([1, 1, 2, 2])
    select_all   = btn1.button("☑ Select all")
    deselect_all = btn2.button("☐ Deselect all")

    if select_all:
        # Streamlit stores the editor key as an internal diff dict, not a DataFrame,
        # so we cannot read it here. Start from val_corrections and set Accept.
        df = st.session_state["val_corrections"].copy()
        df["Accept"] = True
        st.session_state["val_corrections"] = df
        st.session_state["editor_key_version"] = _editor_version + 1
    if deselect_all:
        df = st.session_state["val_corrections"].copy()
        df["Accept"] = False
        st.session_state["val_corrections"] = df
        st.session_state["editor_key_version"] = _editor_version + 1

    # Recompute after possible increment so the correct key is passed below.
    _editor_key = f"corrections_editor_{st.session_state.get('editor_key_version', 0)}"

    # Count accepted rows: start from base corrections, then apply any Accept
    # changes the user has made in the editor (stored as a diff dict, not written
    # back to val_corrections until Apply is clicked). Without this, checking a
    # manual-row Accept box would not enable the button.
    _base = st.session_state["val_corrections"]
    _n_sel = int(_base["Accept"].sum())
    _editor_state = st.session_state.get(_editor_key)
    if isinstance(_editor_state, dict):
        for _idx_str, _changes in _editor_state.get("edited_rows", {}).items():
            if "Accept" in _changes:
                try:
                    _orig = bool(_base.iloc[int(_idx_str)]["Accept"])
                except (IndexError, ValueError):
                    _orig = False
                if _changes["Accept"] and not _orig:
                    _n_sel += 1
                elif not _changes["Accept"] and _orig:
                    _n_sel -= 1

    apply_top = btn3.button("Apply selected fixes", type="primary", disabled=_n_sel == 0)

    # Reserve a slot for the validation warning above the table.
    warning_slot = st.empty()

    # The key lets Streamlit persist the widget state across reruns without
    # reverting to the passed base data on each rerun (no double-click issue).
    # val_corrections is only used as the initial value when the key is absent;
    # after that the key state is the single source of truth for the editor.
    edited = st.data_editor(
        st.session_state["val_corrections"],
        key=_editor_key,
        column_config={
            "Accept":               st.column_config.CheckboxColumn("Accept", default=False),
            "Fix type":             st.column_config.TextColumn("Fix type",             disabled=True),
            "Item ID":              st.column_config.TextColumn("Item ID",              disabled=True),
            "Column":               st.column_config.TextColumn("Column",               disabled=True),
            "Error description":    st.column_config.TextColumn("Error description",    disabled=True),
            "Wrong data":           st.column_config.TextColumn("Wrong data",           disabled=True),
            "Suggested fix":        st.column_config.TextColumn("✏ Suggested fix"),     # editable
            "Description EN (ref)": st.column_config.TextColumn("Description EN (ref)", disabled=True),
        },
        use_container_width=True,
        hide_index=True,
    )

    # Guard: if "Wrong data" was accidentally modified in the editor state
    # (disabled=True is not always enforced by all Streamlit builds), preserve
    # the user's mutable edits and force the editor to re-initialise from the
    # correct base so the original wrong values are restored visually.
    _base = st.session_state["val_corrections"]
    if len(edited) == len(_base) and not edited["Wrong data"].reset_index(drop=True).equals(
        _base["Wrong data"].reset_index(drop=True)
    ):
        _save = _base.copy()
        _save["Accept"]        = edited["Accept"].values
        _save["Suggested fix"] = edited["Suggested fix"].values
        st.session_state["val_corrections"] = _save
        st.session_state["editor_key_version"] = st.session_state.get("editor_key_version", 0) + 1
        st.rerun()

    accepted_rows = edited[edited["Accept"]]
    st.caption(f"{len(accepted_rows)} of {len(edited)} corrections selected")

    # Live validation — runs on every rerun against the current editor state.
    # Checks reference catalogue, uppercase, length, forbidden symbols, spaces.
    _ref_live     = st.session_state["val_ref"]
    _valid_names  = set(_ref_live["basic_name"].dropna().astype(str))
    _valid_groups = set(_ref_live["product_group_id"].astype(str))
    _live_invalid = []
    for _, _row in accepted_rows.iterrows():
        _col = _row["Column"]
        _val = str(_row["Suggested fix"]).strip()
        _ft  = "auto" if "🔧" in str(_row.get("Fix type", "")) else "manual"
        if not _val or _val == "nan":
            continue
        msg = _validate_fix(_row["Item ID"], _col, _val, _valid_names, _valid_groups, _ft)
        if msg:
            _live_invalid.append(msg)

    if _live_invalid:
        warning_slot.warning(
            "⚠️ Some values are invalid — correct them before applying:\n\n"
            + "\n".join(_live_invalid)
        )
    else:
        warning_slot.empty()

    _fixed_applied = st.session_state.get("val_fixed_applied", False)
    if _fixed_applied:
        acc = st.session_state.get("val_accepted_cells", set())
        st.success(
            f"Corrections applied — {len(acc)} cells updated. "
            "Change the selection and apply again to adjust."
        )

    if apply_top:
        # Re-validate all accepted rows before writing.
        _ref          = st.session_state["val_ref"]
        _valid_names  = set(_ref["basic_name"].dropna().astype(str))
        _valid_groups = set(_ref["product_group_id"].astype(str))

        invalid_msgs = []
        for _, row in accepted_rows.iterrows():
            col = row["Column"]
            val = str(row["Suggested fix"]).strip()
            ft  = "auto" if "🔧" in str(row.get("Fix type", "")) else "manual"
            if not val or val == "nan":
                continue
            msg = _validate_fix(row["Item ID"], col, val, _valid_names, _valid_groups, ft)
            if msg:
                invalid_msgs.append(msg)

        if invalid_msgs:
            st.error(
                "Some values failed validation — correct them before applying:\n\n"
                + "\n".join(invalid_msgs)
            )
            return

        # Apply accepted corrections on top of the most recently corrected data,
        # not the original — so successive Apply rounds accumulate fixes rather
        # than restarting from scratch.
        # Convert target column to object dtype first to avoid dtype mismatch
        # (e.g. product_group stored as int64, suggestion is str).
        # Manual rows with an empty Suggested fix are skipped — the cell
        # stays red in the results until the user fills in a value.
        _items      = st.session_state.get("val_fixed_items", st.session_state["val_items"])
        fixed_items = _items.copy()
        accepted_cells = set()
        for _, row in accepted_rows.iterrows():
            col     = row["Column"]
            fix_val = str(row["Suggested fix"]).strip()
            if not fix_val or fix_val == "nan":
                continue
            if col in fixed_items.columns and not pd.api.types.is_object_dtype(fixed_items[col]):
                fixed_items[col] = fixed_items[col].astype(object)
            mask = fixed_items["item_id"] == row["Item ID"]
            fixed_items.loc[mask, col] = fix_val
            accepted_cells.add((row["Item ID"], col))

        # Suggestions that existed but were not accepted → shown yellow in Results
        _all_sugg = st.session_state.get("val_suggestions", pd.DataFrame())
        rejected_cells = (
            set(zip(_all_sugg["item_id"], _all_sugg["column"])) - accepted_cells
            if not _all_sugg.empty else set()
        )

        fixed_results = run_checks(fixed_items, _ref)
        fixed_summ    = summary(fixed_results)

        # Rebuild the corrections table from the now-fixed data so each row
        # reflects the current state — resolved errors drop out and stale
        # "before" values are replaced with the corrected values.
        new_corrections, new_suggestions = _build_corrections(fixed_items, fixed_results, _ref)

        # Increment key version so the editor re-initialises from the rebuilt
        # corrections table rather than keeping the now-stale accepted rows.
        st.session_state["editor_key_version"] = st.session_state.get("editor_key_version", 0) + 1

        st.session_state.update(
            val_fixed_applied=True,
            val_fixed_items=fixed_items,
            val_fixed_results=fixed_results,
            val_fixed_summary=fixed_summ,
            val_accepted_cells=accepted_cells,
            val_rejected_cells=rejected_cells,
            val_corrections=new_corrections,
            val_suggestions=new_suggestions,
            switch_to_results=True,
        )
        st.rerun()


with tab_autofix:
    render_fix_tab()


# ── Tab 3: Download ───────────────────────────────────────────────────────
with tab_download:

    def _add_suggestion_cols(df, suggestions_df):
        """
        Insert a 'Suggested fix: {col}' column right after each data column
        that has at least one suggestion. Only the rows with a suggestion for
        that column get a value; all other cells are empty.
        """
        if suggestions_df is None or suggestions_df.empty:
            return df
        # {column: {item_id: suggested_value}}
        sugg_map = {}
        for _, row in suggestions_df.iterrows():
            sugg_map.setdefault(row["column"], {})[row["item_id"]] = row["after"]

        result = df.copy()
        # Iterate in reverse so earlier insertions don't shift later positions.
        for col in reversed([c for c in df.columns if c in sugg_map]):
            pos = result.columns.get_loc(col) + 1
            values = (
                result["item_id"].map(sugg_map[col])
                if "item_id" in result.columns
                else pd.Series("", index=result.index)
            )
            result.insert(pos, f"Suggested fix: {col}", values)
        return result

    def _apply_excel_colours(ws, df, fixed_set=None, rejected_set=None):
        """
        Colour cells in the Excel results worksheet.

          fixed_set    — (item_id, column) pairs that were accepted → green
          rejected_set — (item_id, column) pairs that had a suggestion but were rejected → yellow
        Suggestion columns (named 'Suggested fix: X') get yellow where non-empty.
        """
        from openpyxl.styles import PatternFill

        pass_fill  = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        fail_fill  = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        warn_fill  = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        fixed_fill = PatternFill(start_color="C3E6CB", end_color="C3E6CB", fill_type="solid")

        col_index  = {name: i + 1 for i, name in enumerate(df.columns)}
        id_col_idx = col_index.get("item_id")

        # Pass 1 — data cells (original columns only, not suggestion columns)
        bad_cells = bad_data_cells(df)
        for col_name, series in bad_cells.items():
            if col_name not in col_index:
                continue
            excel_col = col_index[col_name]
            for row_idx, is_bad in enumerate(series, start=2):
                item_id = df.iloc[row_idx - 2]["item_id"] if id_col_idx else None
                if fixed_set and (item_id, col_name) in fixed_set:
                    ws.cell(row=row_idx, column=excel_col).fill = fixed_fill
                elif rejected_set and (item_id, col_name) in rejected_set:
                    ws.cell(row=row_idx, column=excel_col).fill = warn_fill
                elif is_bad:
                    ws.cell(row=row_idx, column=excel_col).fill = fail_fill

        # Pass 2 — check result columns
        for check_col in CHECK_COLS:
            if check_col not in col_index:
                continue
            excel_col = col_index[check_col]
            for row_idx, val in enumerate(df[check_col], start=2):
                if str(val).strip().lower() == "specification empty":
                    fill = warn_fill
                elif is_pass(check_col, val):
                    fill = pass_fill
                else:
                    fill = fail_fill
                ws.cell(row=row_idx, column=excel_col).fill = fill

        # Pass 3 — suggestion columns: yellow where a suggestion exists
        for col_name, excel_col in col_index.items():
            if not col_name.startswith("Suggested fix: "):
                continue
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=excel_col)
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = warn_fill

    buf = io.BytesIO()
    _suggestions = st.session_state.get("val_suggestions", pd.DataFrame())

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Sheet 1: source data, no check columns
        items.to_excel(writer, sheet_name="original_data", index=False)

        # Sheet 2: validation results — colour-coded, no suggestion columns
        results.to_excel(writer, sheet_name="check_results_with_errors", index=False)
        _apply_excel_colours(writer.sheets["check_results_with_errors"], results)

        # Sheet 3 (if corrections applied): corrected data — green = fixed, yellow = rejected suggestion
        if fixed_applied:
            fixed_res    = st.session_state["val_fixed_results"]
            fixed_set    = st.session_state.get("val_accepted_cells", set())
            rejected_set = st.session_state.get("val_rejected_cells", set())
            # Show suggestions only for cells that were not accepted (still wrong)
            rejected_sugg = _suggestions[
                _suggestions.apply(
                    lambda r: (r["item_id"], r["column"]) in rejected_set, axis=1
                )
            ] if not _suggestions.empty else pd.DataFrame()
            fixed_with_sugg = _add_suggestion_cols(fixed_res, rejected_sugg)
            fixed_with_sugg.to_excel(writer, sheet_name="fixed_data", index=False)
            _apply_excel_colours(
                writer.sheets["fixed_data"], fixed_with_sugg,
                fixed_set=fixed_set, rejected_set=rejected_set,
            )

        # Sheet 4: summary — always last
        (st.session_state["val_fixed_summary"] if fixed_applied else summ).to_excel(
            writer, sheet_name="final_check_summary", index=False
        )

    sheets = "original_data → check_results_with_errors"
    if fixed_applied:
        sheets += " → fixed_data"
    sheets += " → final_check_summary"

    st.caption(
        f"Sheets: **{sheets}**  \n"
        "**original_data** — source data without check columns  \n"
        "**check_results_with_errors** — validation results, colour-coded  \n"
        "**fixed_data** — after corrections: 🟢 green = corrected, 🟡 yellow = suggestion rejected  \n"
        "**final_check_summary** — pass/fail counts per check"
    )
    st.download_button(
        label="Download results (.xlsx)",
        data=buf.getvalue(),
        file_name="validation_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# After Apply selected fixes: switch to Results tab via JS.
# The flag is set in render_fix_tab() before st.rerun() so this block only
# fires on the rerun that follows an Apply — not on every render.
if st.session_state.pop("switch_to_results", False):
    import streamlit.components.v1 as components
    components.html("""
    <script>
    setTimeout(function () {
        var tabs = window.parent.document.querySelectorAll('button[data-baseweb="tab"]');
        if (tabs.length > 0) tabs[0].click();
    }, 150);
    </script>
    """, height=0)
